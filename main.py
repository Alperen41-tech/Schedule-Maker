from openpyxl import load_workbook,Workbook
from openpyxl.styles import PatternFill,Side,Border

def same(list_1,list_2):
    for i in list_1:
        if i in list_2:
            return "var"
        else:
            pass
    for i in list_2:
        if i in list_1:
            return "var"
        else:
            pass

    return "yok"


exen_x = ["A", "B", "C", "D", "E"]
exen_y = ["1","2","3","4","6","7","8","9"]
coordinates = []
for i in exen_x:
    for k in exen_y:
        coordinates.append(i+k)









mbg_sources_list = []
cs_sources_list= []
math_sources_list= []
eng_sources_list = []

file_and_lists_names_list = [("mbg_sources",mbg_sources_list),("cs_sources",cs_sources_list),("math_sources",math_sources_list),("eng_sources",eng_sources_list)]


def sources(file_name,list_name):
    workbook = load_workbook("{}.xlsx".format(file_name))
    worksheet = workbook.active

    alfabeth = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M"]
    len_col = 0
    for i in alfabeth:
        wanted_value = i + "1"
        if worksheet["{}".format(wanted_value)].value != None:
            len_col += 1
        else:
            break

    number = 1

    while number <= len_col:
        liste = []
        for i in range(1, 40):
            value = worksheet["{}{}".format(chr(64 + number), i)].value
            if value != None:
                liste.append(value)
            else:
                break
        list_name.append(liste)
        number += 1



for i,j in file_and_lists_names_list:
    sources(i,j)


def function(differ,list_name):
    for mm in differ:
        list_name.append(mm)



program_number = 1

"""
cs_color = PatternFill(patternType="solid",fgColor="76BA99")
mbg_color = PatternFill(patternType="solid",fgColor="AC4425")
math_color= PatternFill(patternType="solid",fgColor="A0C3D2")
eng_color = PatternFill(patternType="solid",fgColor="E6B325")

type = Side(border_style="thin")
border_type = Border(top=type,right=type,left=type,bottom=type)
"""

cs_color = PatternFill("solid","76BA99")
mbg_color = PatternFill("solid","AC4425")
math_color= PatternFill("solid","A0C3D2")
eng_color = PatternFill("solid","E6B325")

type = Side("thin")
border_type = Border(type,type,type,type)


def extraction(list_1,list_2): #liste 1'den liste 2'nin elamanlarını çıkartır
    for i in list_2:
        if i in list_1:
            list_1.remove(i)
        else:
            pass


def examination(list_1,list_2): #liste 2'nin liste 1'in içinde olup olmadığını sorgular
    for i in list_2:
        if i in list_1:
            pass
        else:
            return False
    return True

def yoksa_ekle(list_1,list_2):
    if examination(list_1,list_2) == False:
        for i in list_2:
            list_1.append(i)




exen_x = ["A", "B", "C", "D", "E"]
exen_y = ["1", "2", "3", "4", "6", "7", "8", "9"]
coordinates = []
for i in exen_x:
    for k in exen_y:
        coordinates.append(i + k)


for i in cs_sources_list:
    actual = []
    for j in math_sources_list:
        actual = []
        function(i, actual)
        if same(j,actual) == "var":
            pass
        else:
            for e in eng_sources_list:
                actual = []
                function(i, actual)
                function(j, actual)
                if same(actual,e) == "var":
                    pass
                else:
                    function(e,actual)
                    for k in mbg_sources_list:
                        actual = []
                        function(i, actual)
                        function(j, actual)
                        function(e,actual)
                        if same(actual, k) == "var":
                            pass
                        else:
                            wb = Workbook()
                            ws = wb.active
                            for mm in i[1:]:
                                ws[mm].fill = cs_color
                            for mn in j[1:]:
                                ws[mn].fill = math_color
                            for jj in e[1:]:
                                ws[jj].fill = eng_color
                            for kk in k[1:]:
                                ws[kk].fill = mbg_color

                            for nn in coordinates:
                                ws[nn].border = border_type

                            liste_name =[i[0],j[0],e[0],k[0]]

                            for t in range(0,4):
                                ws["G"+str(t+2)].value = liste_name[t]

                            ws.column_dimensions["G"].width = 20


                            wb.save("D:/Python/projeler/courses/programs/program_{}.xlsx".format(program_number))
                            program_number += 1
print(program_number-1)






"""
coordinates_main = coordinates



for i in cs_sources_list:

    coordinates = coordinates_main
    print(coordinates)
    extraction(coordinates,i[1:])
    print(coordinates)
    for j in math_sources_list:
        if examination(coordinates,j[1:]) == False:
            print("false döndü1")
            pass
        else:
            print("true döndü2")
            extraction(coordinates,j)
            print(coordinates)
            for k in eng_sources_list:
                if examination(coordinates,k[1:]) == False:
                    print("false döndü3")
                    pass
                else:
                    print("true döndü4")
                    extraction(coordinates,k)
                    print(coordinates)
                    for l in mbg_sources_list:
                        if examination(coordinates,l[1:]) == False:
                            print("false döndü5")
                            pass
                        else:
                            print("true döndü6")
                            print(coordinates)
                            wb = Workbook()
                            ws = wb.active
                            for mm in i[1:]:
                                ws[mm].fill = cs_color
                            for mn in j[1:]:
                                ws[mn].fill = math_color
                            for jj in k[1:]:
                                ws[jj].fill = eng_color
                            for kk in l[1:]:
                                ws[kk].fill = mbg_color

                            for nn in coordinates_main:
                                ws[nn].border = border_type

                            liste_name = [i[0], j[0], k[0], l[0]]

                            for t in range(0, 4):
                                ws["G" + str(t + 2)].value = liste_name[t]

                            ws.column_dimensions["G"].width = 20

                            wb.save("D:/Python/projeler/courses/programs/program_{}.xlsx".format(program_number))
                            program_number += 1
"""



















